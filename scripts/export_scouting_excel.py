# -*- coding: utf-8 -*-
"""Izvoz scouting tablice iz Supabase u Excel (31.07.2026).

Supabase je izvor istine (ondje su ispravci i 50 novih profila); ova skripta iz njega
generira Excel istog oblika kao korisnikov original, da se tablica moze citati i uredjivati
rucno. Stari "top100" file se NE dira — ostaje kao zapis polaznog stanja.

Ako korisnik rucno uredi izvezeni Excel, `scripts/import_scouting.py` ga vraca natrag u bazu
(upsert po player_name), pa krug ostaje zatvoren u oba smjera.

Pokretanje:  python scripts/export_scouting_excel.py
"""
import sys
import os
import io

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database import supabase_client as db

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..",
                   "ATP_Player_Scouting top150.xlsx")
SHEET = "ATP Player Scouting"

HEADERS = ["Rank", "Player / Igrac", "Country", "Hand / BH", "Style / Stil",
           "Best surfaces / Najbolje podloge", "Strengths / Prednosti",
           "Weaknesses / Mane", "Favourable matchups / Voli igrati protiv",
           "Tough matchups / Muku muci s", "Catchy note / Zanimljivost",
           "Confidence / Pouzdanost"]
WIDTHS = [7, 26, 9, 16, 34, 30, 46, 46, 42, 42, 70, 14]
FIELDS = ["rank", "display_name", "country", "hand", "style", "best_surfaces",
          "strengths", "weaknesses", "favourable_matchups", "tough_matchups",
          "note", "confidence"]

LEGEND = [
    "LEGENDA / LEGEND",
    "Confidence: High > Med-High > Med > Med-Low > Low > Insufficient. "
    "Model koristi profil SAMO od Med-Low navise; Low i Insufficient se ignoriraju.",
    "Profil utjece na procjenu najvise +/-3pp i sluzi kao tie-breaker kad su mjereni "
    "faktori izjednaceni — nikada ne moze nadjacati ELO, formu ili statistiku servisa.",
    "Profili rangova ~100-150 (dodani 31.07.2026) IZVEDENI su iz mjerenih podataka "
    "(3-godisnji ucinak po podlogama, hold%, return points won, ace rate, surface ELO), "
    "a ne iz neovisnog skautiranja. Gdje podatka nema, pise 'cannot be determined'.",
    "Za igrace ranga >100 trogodisnji W-L ukljucuje i Challenger/ITF razinu, pa je "
    "stvarna kvaliteta na ATP razini obicno niza od sirovog postotka.",
    "Ako se vanjski opis i izmjerene brojke ne slazu, u biljesci stoji CONFLICT i "
    "uputa da se vjeruje izmjerenom (primjer: Arthur Gea).",
    "Uredjivanje: promijeni sto zelis pa pokreni  python scripts/import_scouting.py  "
    "da se vrati u Supabase. Supabase je izvor istine za dnevni model.",
]


def main():
    rows = db._select("player_scouting", select="*", limit=400)
    rows.sort(key=lambda r: (r.get("rank") or 9999, (r.get("display_name") or "")))
    print(f"Dohvaceno profila iz Supabase: {len(rows)}")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET

    head_fill = PatternFill("solid", fgColor="1F4E78")
    head_font = Font(bold=True, color="FFFFFF", size=11)
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = WIDTHS[c - 1]
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "C2"

    # Blaga vizualna oznaka pouzdanosti — da se na prvi pogled vidi cemu se moze vjerovati
    conf_fill = {
        "High":         PatternFill("solid", fgColor="C6EFCE"),
        "Med-High":     PatternFill("solid", fgColor="DDF3DC"),
        "Med":          PatternFill("solid", fgColor="FFF2CC"),
        "Med-Low":      PatternFill("solid", fgColor="FCE4D6"),
        "Low":          PatternFill("solid", fgColor="F2F2F2"),
        "Insufficient": PatternFill("solid", fgColor="E7E6E6"),
    }

    for i, r in enumerate(rows, start=2):
        for c, f in enumerate(FIELDS, 1):
            v = r.get(f)
            cell = ws.cell(row=i, column=c, value=v if v is not None else "")
            cell.alignment = Alignment(vertical="top", wrap_text=(c >= 5))
        conf = (r.get("confidence") or "").strip()
        if conf in conf_fill:
            ws.cell(row=i, column=12).fill = conf_fill[conf]

    start = len(rows) + 3
    ws.cell(row=start, column=1, value=LEGEND[0]).font = Font(bold=True)
    for j, line in enumerate(LEGEND[1:], 1):
        ws.cell(row=start + j, column=1, value=line).alignment = Alignment(wrap_text=False)

    wb.save(OUT)
    by_conf = {}
    for r in rows:
        by_conf[r.get("confidence")] = by_conf.get(r.get("confidence"), 0) + 1
    print(f"Spremljeno: {os.path.basename(OUT)}")
    print(f"  redaka igraca: {len(rows)} (rank {rows[0].get('rank')}..{rows[-1].get('rank')})")
    print(f"  po confidence: {dict(sorted(by_conf.items(), key=lambda x: -x[1]))}")


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    main()
